import PySimpleGUI as sg
import os
from pathlib import Path
import shutil

import win32com.client

from functions import convertCADtoOSGB36, importSnakeGrid, pasteProfile,get3DCoordinates

editingSheetPath = ""
editingProjectPath = ""


def setup_shortcut():
    driveName = "OneDrive - SystraGroup"
    desktopPath = str(Path.home() / driveName / "Desktop")
    folderPath = f"{desktopPath}\CAD2OpenRoads"
    shortcutPath = f"{folderPath}\CAD2OpenRoads - Atalho.lnk"

    shell = win32com.client.Dispatch("WScript.Shell")

    if not os.path.exists(folderPath):
        resourcesName = "resources"
        os.mkdir(folderPath)
        os.mkdir(f"{folderPath}\Projects")
        os.mkdir(f"{folderPath}\{resourcesName}")
        shutil.copyfile(os.path.abspath('resources/CoordinatesBaseSheet.xlsx'), f"{folderPath}\{resourcesName}\CoordinatesBaseSheet.xlsx")

    if os.path.exists("main.exe"):
        try:
            shortcut = shell.CreateShortCut(shortcutPath)
            shortcut.Targetpath = os.path.abspath("main.exe")
            shortcut.WindowStyle = 7
            shortcut.save()
        except:
            print("Erro! Desktop não encontrado!")


def open_menu():
    global editingSheetPath
    editingSheetPath = ""

    layoutMenu = [
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('Gerador de Planilhas de Coordenadas', font=('Arial', 15))],
        [
            sg.Text('AutoCAD para OpenRoads')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Criar novo arquivo', size=(20,5), key='ButtonCreateNewFIle'),
            sg.Button('Editar arquivo', size=(20,5), key='ButtonEditFile')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Fechar', size=(15, 3), key='ButtonClose'),
            sg.Button('⚙️', font=('Arial', 14), size=(5, 2), key='ButtonSettings')],
    ]

    window = sg.Window('Gerador de Planilhas das Coordenadas', layoutMenu, element_justification='c', size=(500, 350))

    while True:
        event, values = window.Read()

        if event == 'ButtonCreateNewFIle':
            window.close()
            open_new_sheet_window()

        if event == 'ButtonEditFile':
            window.close()
            open_edit_sheet_window()

        if event == "ButtonSettings":
            setup_shortcut()

        if event == sg.WINDOW_CLOSED or event == "ButtonClose":
            break

    window.close()


def open_new_sheet_window():
    global editingSheetPath
    global editingProjectPath

    layoutMenu = [
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('Criar novo arquivo', font=('Arial', 15))],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('Dê um nome para o projeto: ', font=('Arial', 13))],
        [
            sg.InputText('MyCoordinatesSheet', pad=(10, 10), size=(25, 5), font=('Arial', 13), justification='centralized', key='InputSheetName')],
        [
            sg.Text('', pad=(0, 10))],
        [
            sg.Text('Escolha uma pasta para salvar o projeto: ', font=('Arial', 13))],
        [
            sg.InputText(os.path.abspath('Projects'), size=(45, 5), font=('Arial', 10), pad=(10, 20), key='InputSheetPath'),
            sg.FolderBrowse('Procurar', initial_folder=os.path.abspath('Projects'))],
        [
            sg.Text('',  visible=False, key='TextCreate')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Criar arquivo', size=(25, 0), key='ButtonCreate')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Voltar', size=(15, 1), key='ButtonReturn')],
    ]


    windowNewSheet = sg.Window('Gerador de Planilhas das Coordenadas', layoutMenu, element_justification='c', size=(500, 500))

    while True:
        event, values = windowNewSheet.Read()
        #"{values['InputSheetName']}.xlsx"
        if event == "ButtonCreate":
            if os.path.exists(f"{values['InputSheetPath']}") == False:
                os.mkdir(f"{values['InputSheetPath']}")

            folderPath = f"{values['InputSheetPath']}/{values['InputSheetName']}"
            if not os.path.exists(folderPath):
                os.mkdir(folderPath)
                shutil.copyfile(os.path.abspath('resources/CoordinatesBaseSheet.xlsx'), f"{folderPath}/{values['InputSheetName']}.xlsx")
                windowNewSheet['TextCreate'].update(visible=True)
                windowNewSheet['TextCreate'].update("Planilha criada com sucesso!")
                editingProjectPath = folderPath
                editingSheetPath = f"{folderPath}\{values['InputSheetName']}.xlsx"
                windowNewSheet.close()
                open_steps_window()

            else:
                windowNewSheet['TextCreate'].update(visible=True)
                windowNewSheet['TextCreate'].update("Já existe uma pasta com este nome nesse local. Tente novamente!")


        if event == "ButtonReturn":
            windowNewSheet.close()
            open_menu()

        if event == sg.WINDOW_CLOSED:
            break

    windowNewSheet.close()


def open_edit_sheet_window():
    global editingSheetPath
    global editingProjectPath

    layoutEditSheet = [
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('Editar arquivo existente', font=('Arial', 15))],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('Escolha um arquivo .xlsx para editar: ', font=('Arial', 13))],
        [
            sg.InputText(size=(35, 5), font=('Arial', 10), pad=(10, 20), key='InputSheetPath'),
            sg.FileBrowse('Procurar', initial_folder=os.path.abspath('Projects'), file_types=(("Microsoft Excel workbook after Excel 2007 Files", "*.xlsx"),))],
        [
            sg.Text('', visible=False, key='TextOpen')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Abrir arquivo', size=(25, 0), key='ButtonOpen')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Voltar', size=(15, 1), key='ButtonReturn')],
    ]

    windowEditSheet = sg.Window('Gerador de Planilhas das Coordenadas', layoutEditSheet, element_justification='c',
                               size=(500, 400))

    while True:
        event, values = windowEditSheet.Read()

        #validação do tipo do arquivo
        if event == "ButtonOpen":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=values['InputSheetPath'])
                ws = wb.active
                templateCode = ws['A1'].value
                wb.close()

                if templateCode == 'CoordinatesSheetGeneratorFile':
                    windowEditSheet['TextOpen'].update(visible=True)
                    windowEditSheet['TextOpen'].update("Abrindo arquivo!")
                    editingProjectPath = os.path.dirname(values['InputSheetPath'])
                    editingSheetPath = f"{values['InputSheetPath']}"
                    windowEditSheet.close()
                    open_steps_window()

                else:
                    windowEditSheet['TextOpen'].update(visible=True)
                    windowEditSheet['TextOpen'].update("Esse arquivo não está no formato padrão do programa.\nTente novamente com outro arquivo!")
            except:
                windowEditSheet['TextOpen'].update(visible=True)
                windowEditSheet['TextOpen'].update("Formato do arquivo inválido! Formato padrão: .xlsx")

        if event == "ButtonReturn":
            windowEditSheet.close()
            open_menu()

        if event == sg.WINDOW_CLOSED:
            break

    windowEditSheet.close()


def open_steps_window():

    layoutStepsWindow = [
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('Utilize as funções abaixo:', font=('Arial', 15))],

        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Colar LIST 2D', size=(40, 2), font=('Arial', 15), key='ButtonPaste2DList')],
        [
            sg.Button('Importar SnakeGrid', size=(40, 2),  font=('Arial', 15), key='ButtonImportSnakeGrid')],
        [
            sg.Button('Colar LIST Perfis', size=(40, 2),  font=('Arial', 15), key='ButtonPasteProfileList')],
        [
            sg.Button('Obter Coordenadas 3D',  font=('Arial', 15), size=(40, 2), key='ButtonORProfileConverter')],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Text('', pad=(0, 5), visible=False, key="TextAlertSteps")],
        [
            sg.Text('', pad=(0, 5))],
        [
            sg.Button('Voltar', size=(15, 1), key='ButtonReturn')]
    ]

    windowSteps = sg.Window('Gerador de Planilhas das Coordenadas', layoutStepsWindow, element_justification='c', size=(500, 550))

    while True:
        event, values = windowSteps.Read()

        if event == 'ButtonPaste2DList':
            convertCADtoOSGB36(editingSheetPath, windowSteps)

        if event == 'ButtonImportSnakeGrid':
            fileFormat = ["CSV Files", "*.csv"]
            snakeGridSheetPath = open_select_file_window(fileFormat)
            if str(snakeGridSheetPath).endswith(".csv"):
                importSnakeGrid(editingSheetPath, editingProjectPath, snakeGridSheetPath, windowSteps)
            else:
                windowSteps['TextAlertSteps'].update(visible=True)
                windowSteps['TextAlertSteps'].update("Formato de arquivo incompatível! Tente novamente!")

        if event == 'ButtonPasteProfileList':
            pasteProfile(editingSheetPath, editingProjectPath, windowSteps)

        if event == 'ButtonORProfileConverter':
            get3DCoordinates(editingSheetPath, editingProjectPath, windowSteps)

        if event == "ButtonReturn":
            windowSteps.close()
            open_menu()

        if event == sg.WINDOW_CLOSED:
            break

    windowSteps.close()


def open_select_file_window(fileFormat):
    layoutSelectWindow = [
        [
            sg.Text('', pad=(0, 1))],
        [
            sg.Text('Escolha um arquivo: ')],
        [
            sg.InputText(size=(35, 5), font=('Arial', 10), pad=(10, 20), key='InputSheetPath'),
            sg.FileBrowse('Procurar', initial_folder=str(Path.home() / "Downloads"), file_types=((fileFormat[0], fileFormat[1]),))],
        [
            sg.Button('Importar arquivo', size=(25, 0), key='ButtonImport')]
    ]

    windowFileSelect = sg.Window('Gerador de Planilhas das Coordenadas', layoutSelectWindow, element_justification='c',
                                size=(500, 175))

    while True:
        event, values = windowFileSelect.Read()

        if event == 'ButtonImport':
            sheetPath = values['InputSheetPath']
            windowFileSelect.close()
            return sheetPath

        if event == sg.WINDOW_CLOSED or event == "ButtonClose":
            break

    windowFileSelect.close()

def main():
    sg.theme('DarkRed')
    open_menu()


main()