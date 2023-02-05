import numpy as np
import pandas as pd
import win32clipboard

# converter LIST copiada do AutoCAD para formato OSGB36


def limparColuna(column, ws, cellType):
    for n in range(1, len(ws[column])+1):
        cell = ws[f'{column}{n}']
        if n >= 6 and type(cell.value) == cellType:
            cell.value = None


def convertCADtoOSGB36(editingSheetPath, window):

    try:
        win32clipboard.OpenClipboard()
        listValue = win32clipboard.GetClipboardData()
        win32clipboard.CloseClipboard()

        # Abrir arquivo excel para salvar alterações
        from openpyxl import load_workbook
        editingSheetPath = editingSheetPath.replace("/", "\\")
        wb = load_workbook(filename=editingSheetPath)
        ws = wb.active

        listValue = " at point" + listValue.split('at point', 1)[1]
        listValue = listValue.replace(' Press ENTER to continue:', '')  # Remover ' Press ENTER to continue:'
        listValue = listValue.replace('\r\n         ', '') # Remover quebras de linha
        listValue = listValue.split(" at point  ")  # Separar o texto a cada ' at point  '
        listValue.pop(0) # Remover primeiro elemento (vazio)

        wb.save(editingSheetPath)

        # Limpar valores das coordenadas dos 2D (colunas B e C)
        limparColuna('B', ws, str)
        limparColuna('C', ws, str)

        clipboardCoordinates = ""

        #Fazer tratamento dos dados do List (AutoCAD)
        for n in range(len(listValue)):
            coord = listValue[n].split("=")  # Separa o texto de cada coordenada a cada '=  '
            x = float(coord[1].replace("Y", ""))  # Substitui o Y da coordenada por vazio "" e transforma em float
            y = float(coord[2].replace("Z", ""))  # Substitui o Z da coordenada por vazio "" e transforma em float
            x = str(x).replace('.', ',')  # Substitui ponto por virgula
            y = str(y).replace('.', ',')  # Substitui ponto por virgula

            #Alterar respectivas células na planilha
            ws[f'B{n+6}'].value = x
            ws[f'C{n+6}'].value = y

            #Adicionar ao texto na área de transferência
            clipboardCoordinates += f"{x}    {y} "

        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(clipboardCoordinates, win32clipboard.CF_TEXT)
        win32clipboard.CloseClipboard()

        wb.save(editingSheetPath)
        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update("(Passo 1 - Completo)\nInformações salvas com sucesso!\nCoordenadas copiadas para a área de transferência")


    except:
        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update("(Erro - Passo 1)\nVerifique se o texto na área de transferência é válido. \nCertifique-se também de que o arquivo não esteja aberto!")


def importSnakeGrid(editingSheetPath, editingProjectPath, snakeGridSheetPath, window):
    try:
        # Abrir csv das coordenadas SnakeGrid fazendo filtragem das colunas relevantes
        df = pd.read_csv(snakeGridSheetPath, sep='[:, |]', usecols=['Converted_X', 'Converted_Y'], engine='python')

        # Abrir planilha com as informações do projeto
        from openpyxl import load_workbook
        editingSheetPath = editingSheetPath.replace("/", "\\")
        wb = load_workbook(filename=editingSheetPath)
        ws = wb.active

        # Limpar valores das coordenadas 2D Snakegrid (colunas D, E, F, G e M)
        limparColuna('D', ws, float)
        limparColuna('E', ws, float)
        limparColuna('F', ws, float)
        limparColuna('G', ws, float)
        limparColuna('M', ws, str)

        for n in df.index:
            #Inserir coordenadas snakegrid X e Y
            x = df["Converted_X"][n]
            y = df["Converted_Y"][n]
            ws[f'D{n + 6}'].value = x
            ws[f'E{n + 6}'].value = y

            #Calcular e inserir distâncias entre os pontos e somatória delas
            if n != 0:
                dist = np.sqrt((x-ws[f'D{n-1 + 6}'].value)**2 + (y-ws[f'E{n-1 + 6}'].value)**2)
                ws[f'F{n + 6}'].value = dist
                ws[f'G{n + 6}'].value = ws[f'G{n-1 + 6}'].value + dist
            else:
                dist = float(0)
                ws[f'F{n + 6}'].value = 0
                ws[f'G{n + 6}'].value = 0

            #Paronizar no formato do OpenRoads
            ws[f'M{n + 6}'].value = f"xy={x},{y};"


        wb.save(editingSheetPath)

        #Criar .txt no formato do OpenRoads
        openRoadsTxt = "Place smartline\n"
        i=1
        for cell in ws['M']:
            if i>=6:
                openRoadsTxt += f"\n{cell.value}"
            i+=1
        openRoadsTxt += "\n\nReset"

        with open(f"{editingProjectPath}/2D Coordinates OpenRoads.txt", "w") as txt2DCoord:
            txt2DCoord.write(openRoadsTxt)

        clipboardOpenRoads = f'@"{txt2DCoord.name}"'
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(clipboardOpenRoads, win32clipboard.CF_TEXT)
        win32clipboard.CloseClipboard()

        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update("(Passo 2 - Completo)\nArquivo importado com sucesso!\nComando das Coord. 2D do OpenRoads copiado para área de transferência!")

    except:
        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update("(Erro - Passo 2)\nVerifique se o arquivo selecionado é válido. \nCertifique-se também de que o arquivo não esteja aberto!")

def pasteProfile(editingSheetPath, editingProjectPath, window):
    try:
        # Copiar valor da área de transferência
        win32clipboard.OpenClipboard()
        listValue = win32clipboard.GetClipboardData()
        win32clipboard.CloseClipboard()

        # Abrir arquivo excel para salvar alterações
        from openpyxl import load_workbook
        editingSheetPath = editingSheetPath.replace("/", "\\")
        wb = load_workbook(filename=editingSheetPath)
        ws = wb.active

        #Fazer tratamento dos dados do List (AutoCAD)
        listValue = " at point" + listValue.split('at point', 1)[1]
        listValue = listValue.replace(' Press ENTER to continue:', '')  # Remover ' Press ENTER to continue:'
        listValue = listValue.replace('\r\n         ', '') # Remover quebras de linha
        listValue = listValue.split(" at point  ")  # Separar o texto a cada ' at point  '
        listValue.pop(0) # Remover primeiro elemento (vazio)

        clipboardCoordinates = ""

        #Limpar valores das coordenadas dos perfis (colunas Q e R)
        limparColuna('Q', ws, float)
        limparColuna('R', ws, float)
        limparColuna('N', ws, str)

        #Preencher valores das coordenadas dos perfis
        valorInicial = 0
        for n in range(len(listValue)):
            coord = listValue[n].split("=")  # Separa o texto de cada coordenada a cada '=  '
            x = float(coord[1].replace("Y", ""))  # Substitui o Y da coordenada por vazio "" e transforma em float
            y = float(coord[2].replace("Z", ""))  # Substitui o Z da coordenada por vazio "" e transforma em float

            if n == 0:
                valorInicial = x

            x -= float(valorInicial)

            #Alterar respectivas células na planilha
            ws[f'Q{n+6}'].value = x
            ws[f'R{n+6}'].value = y

            #Adicionar ao texto na área de transferência
            clipboardCoordinates += f"{x}    {y} "
            # Coordenadas Perfis OpenRoads
            ws[f'N{n + 6}'].value = f"xy={x},{y};"

        #Salvar valores das coordenadas dos perfis na planilha
        wb.save(editingSheetPath)

        # Criar .txt no formato do OpenRoads
        openRoadsTxt = "Place smartline\n"
        i = 1
        for cell in ws['N']:
            if i >= 6:
                openRoadsTxt += f"\n{cell.value}"
            i += 1
        openRoadsTxt += "\n\nReset"

        with open(f"{editingProjectPath}/Profile Coordinates OpenRoads.txt", "w") as txtProfileCoord:
            txtProfileCoord.write(openRoadsTxt)

        clipboardOpenRoads = f'@"{txtProfileCoord.name}"'
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(clipboardOpenRoads, win32clipboard.CF_TEXT)
        win32clipboard.CloseClipboard()

        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update("(Passo 3 - Completo)\nInformações salvas com sucesso!\nCoordenadas copiadas para a área de transferência")


    except:
        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update(
            "(Erro - Passo 3)\nVerifique se o texto na área de transferência é válido. \nCertifique-se também de que o arquivo não esteja aberto!")


def get3DCoordinates(editingSheetPath, editingProjectPath, window):
    try:
        # Abrir arquivo excel para salvar alterações
        from openpyxl import load_workbook
        editingSheetPath = editingSheetPath.replace("/", "\\")
        wb = load_workbook(filename=editingSheetPath)
        ws = wb.active

        limparColuna('K', ws, float)
        limparColuna('O', ws, str)

        # Calcular Z para cada coordenada 2D
        for n in range(1, len(ws['G'])):
            lenght2D_cell = ws[f'G{n}'].value
            if type(lenght2D_cell) == float or type(lenght2D_cell) == int:
                thisValueIndex = 0
                lastValueIndex = 0
                for m in range(1, len(ws['C'])):
                    lenghtProfile_cell = ws[f'Q{m}'].value
                    if type(lenghtProfile_cell) == float or type(lenghtProfile_cell) == int:
                        thisValueIndex = m
                        if lenghtProfile_cell > lenght2D_cell:
                            break
                        elif lenghtProfile_cell == lenght2D_cell:
                            lastValueIndex = m
                            break
                        else:
                            lastValueIndex = m

                if thisValueIndex != lastValueIndex:
                    alturaPonto = ws[f"R{lastValueIndex}"].value + (
                            (ws[f"G{n}"].value - ws[f"Q{lastValueIndex}"].value) * (
                            ws[f"R{thisValueIndex}"].value - ws[f"R{lastValueIndex}"].value) / (
                                    ws[f"Q{thisValueIndex}"].value - ws[f"Q{lastValueIndex}"].value))
                    ws[f"K{n}"].value = alturaPonto
                    # Coordenadas Perfis OpenRoads
                    ws[f'O{n}'].value = f"xy={ws[f'D{n}'].value},{ws[f'E{n}'].value},{alturaPonto};"

                elif ws[f"G{n}"].value > ws[f'Q{len(ws["Q"])}'].value:
                    Lfinal = ws[f'R{len(ws["R"])}'].value
                    ws[f"K{n}"].value = ws[f'R{len(ws["R"])}'].value
                    ws[f'O{n}'].value = f"xy={ws[f'D{n}'].value},{ws[f'E{n}'].value},{Lfinal};"

                else:
                    ws[f"K{n}"].value = ws[f"R{thisValueIndex}"].value
                    ws[f'O{n}'].value = f"xy={ws[f'D{n}'].value},{ws[f'E{n}'].value},{ws[f'R{thisValueIndex}'].value};"

        #Salvar planilha
        wb.save(editingSheetPath)

        # Criar .txt no formato do OpenRoads
        openRoadsTxt = "Place smartline\n"
        i = 1
        for cell in ws['O']:
            if i >= 6:
                openRoadsTxt += f"\n{cell.value}"
            i += 1
        openRoadsTxt += "\n\nReset"

        with open(f"{editingProjectPath}/3D Coordinates OpenRoads.txt", "w") as txt3DCoord:
            txt3DCoord.write(openRoadsTxt)

        clipboardOpenRoads = f'@"{txt3DCoord.name}"'
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(clipboardOpenRoads, win32clipboard.CF_TEXT)
        win32clipboard.CloseClipboard()

        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update(
            "(Passo 4 - Completo)\nInformações salvas com sucesso!\nCoordenadas copiadas para a área de transferência")

    except:
        window['TextAlertSteps'].update(visible=True)
        window['TextAlertSteps'].update(
            "(Erro - Passo 4)\nVerifique se o texto na área de transferência é válido. \nCertifique-se também de que o arquivo não esteja aberto!")


